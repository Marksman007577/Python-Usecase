import openpyxl

# Opening an Excel Workbook
wb = openpyxl.load_workbook('Advertising.xlsx')

# Getting a particular sheet
adverting_sheet = wb['Advertising']
print(adverting_sheet)

# Get a single cell value using workbook index
single_cell_value = adverting_sheet['A1'].value
print(single_cell_value)

# Get a single cell value using workbook index
a_cell_value = adverting_sheet.cell(row=1, column=2).value
print(a_cell_value)

# Getting the cell method with for loop
cell_list = []
for i in range(1, adverting_sheet.max_row):
    cell_list.append(adverting_sheet.cell(row=i, column=1).value)
print(cell_list)

for i in range(1, adverting_sheet.max_row):
    print(i, adverting_sheet.cell(row=i, column=2).value)

# Finding the highest row/column or lowest row/column
lowest_col = adverting_sheet.min_column
highest_col = adverting_sheet.max_column
lowest_row = adverting_sheet.min_row
highest_row = adverting_sheet.max_row
print('lowest row is', lowest_row)
print('highest row is', highest_row)
print('lowest col is', lowest_col)
print('highest col is', highest_col)