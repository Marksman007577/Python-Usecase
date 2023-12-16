import openpyxl

# Creating an empty workbook a.k.a Excel file
wb = openpyxl.Workbook()

# Visualize some parameters and attributes of the workbook
print(wb.sheetnames)
Sheet = wb.active
print(Sheet)
Sheet.title = "State Population"
State_Population = wb.active

# Add new sheets with their title
wb.create_sheet(title='Year')  # creates a single sheet to the workbook
wb.create_sheet(index=2, title='Population')

# Saves the workbook with a title
wb.save('state_population.xlsx')

# Writing a value to the cell
State_Population['A1'] = 'YEAR'
State_Population['B1'] = 'POPULATION'
State_Population['C1'] = 'SEX'
wb.save('state_population.xlsx')
