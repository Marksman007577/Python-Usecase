import openpyxl

# Reading the Excel Sheet
wb = openpyxl.load_workbook('Advertising.xlsx')
advertising_sheet = wb['Advertising']
print(tuple(advertising_sheet.rows))
print(tuple(advertising_sheet.columns))
print(tuple(advertising_sheet[advertising_sheet.min_column:advertising_sheet.max_column]))
print(tuple(advertising_sheet[advertising_sheet.min_row:advertising_sheet.max_row]))

# Slicing the Whole Excel Sheet
for row in advertising_sheet[advertising_sheet.min_row:advertising_sheet.max_row]:
    for cell in row:
        print(cell.coordinate, cell.value)
    print('----END OF ROW----')

print('New Result starts on next line')
# Getting a slice of the Excel Sheet
for row_x in advertising_sheet['A1':'D15']:
    for cell in row_x:
        print(cell.coordinate, cell.value)
    print('----END OF ROW----')

result_list = []
for nth_row in advertising_sheet['A1':'D15']:
    for cell in nth_row:
        result_list.append(cell.value)
    print('----END OF ROW----')

for i in result_list:
    print(i)
