#!python3

import openpyxl
from openpyxl.styles import Font
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}
FONT_TYPE = Font(name='Times New Roman', size=14, italic=False)

# Opening an Excel file
wb = openpyxl.load_workbook('product_information.xlsx')

# Select worksheet to work with
working_sheet = wb['Product Information']

# Update Worksheet
for row_num in range(2, working_sheet.max_row):
    product_name = working_sheet.cell(row=row_num, column=1).value
    if product_name in PRICE_UPDATES:
        working_sheet.cell(row=row_num, column=1).font = FONT_TYPE
        working_sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[product_name]
wb.save('product_information.xlsx')
